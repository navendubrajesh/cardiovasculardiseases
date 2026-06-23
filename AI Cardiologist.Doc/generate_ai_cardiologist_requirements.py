# -*- coding: utf-8 -*-
"""
Generate AI Cardiologist Requirement.xlsx — product backlog for the AI Cardiologist
research-to-product transition (Big 4 consulting / solution architecture view).

Sources consolidated:
  - MS Navendu Brajesh V1.2.pdf (master's thesis, Aug 2022)
  - AI_Cardiologist_for_journals.pdf (journal submission draft)
  - AI Cardiologist Code/*.ipynb (reference ML implementations)
  - Review Comments 01.txt (peer review remediation)
  - FrenchCoachPro_Requirements.xlsx (backlog workbook structure)
  - AI Cardiologist Working.xlsx (reference bibliography)
  - ResearchPaper Data/ (BRFSS 2015 + LLCP 2024 CDC assets)

Build order (per product owner):
  Phase 1 — ML engine: train, evaluate, export all classifiers + ensemble + explainability
  Phase 2 — Website: prediction UI, batch testing, model comparison dashboards
  Phase 3 — Research hub, GitHub Pages deployment, review-gap hardening
"""
from __future__ import annotations

from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
OUT_XLSX = ROOT / "AI Cardiologist Requirement.xlsx"

REPO_ML = "ML Engine & Training Pipeline"
REPO_WEB = "Website (GitHub Pages)"
REPO_DOCS = "Research & Documentation Hub"

HEADER = [
    "ID", "Repository", "Epic", "Functionality", "User Story", "Role", "Type",
    "Priority", "Phase", "Source", "Acceptance Criteria", "Status", "Research Trace",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
DONE_FILL = PatternFill("solid", fgColor="C6EFCE")
TODO_FILL = PatternFill("solid", fgColor="FFEB9C")
DEFER_FILL = PatternFill("solid", fgColor="D9D9D9")

# ---------------------------------------------------------------------------
# Backlog rows — Status ToDo unless noted; implementation starts from scratch
# ---------------------------------------------------------------------------
BACKLOG: list[dict] = [
    # --- E01 Governance & product framing ---
    {
        "ID": "AC-001",
        "Repository": REPO_DOCS,
        "Epic": "E01 Product Vision & Clinical Governance",
        "Functionality": "Decision-support disclaimer & ethics banner",
        "User Story": "As a visitor, I want a persistent disclaimer that the AI Cardiologist is population-level screening / decision-support only (not a clinical diagnosis), so that misuse and liability risk are mitigated.",
        "Role": "Patient / Clinician",
        "Type": "New",
        "Priority": "Must",
        "Phase": "MVP",
        "Source": "Journal §I + Review Comments (clinical utility)",
        "Acceptance Criteria": "Site-wide banner + /about/governance page citing FUTURE-AI, FAIR-AI, AHA scientific statement; blocks prediction until acknowledged once per session.",
        "Status": "ToDo",
        "Research Trace": "Thesis 1.3, 1.7; Journal Limitations §VI-A",
    },
    {
        "ID": "AC-002",
        "Repository": REPO_DOCS,
        "Epic": "E01 Product Vision & Clinical Governance",
        "Functionality": "Research provenance & citation index",
        "User Story": "As a researcher, I want a searchable bibliography synced to Current_Document_References_REF.txt and AI Cardiologist Working.xlsx, so that every product claim traces to thesis/journal sources.",
        "Role": "Researcher",
        "Type": "New",
        "Priority": "Must",
        "Phase": "MVP",
        "Source": "AI Cardiologist Working.xlsx; Reference/",
        "Acceptance Criteria": "References page lists REF_2–REF_52 with DOI links; filters by topic (XAI, BRFSS, ensembles, ethics); export BibTeX from Refrence.BIB.",
        "Status": "ToDo",
        "Research Trace": "Current_Document_References_REF.txt",
    },
    # --- E02 Data ---
    {
        "ID": "AC-010",
        "Repository": REPO_ML,
        "Epic": "E02 BRFSS Data Pipeline",
        "Functionality": "BRFSS 2015 primary training dataset",
        "User Story": "As the ML engineer, I want a reproducible ingest of heart_disease_health_indicators_BRFSS2015.csv (253,680 rows, 21 features + HeartDiseaseorAttack target), so that notebook results can be replicated in the product pipeline.",
        "Role": "ML Engineer",
        "Type": "Existing",
        "Priority": "Must",
        "Phase": "Phase 1 — ML First",
        "Source": "Thesis §4.2; CVD_*.ipynb",
        "Acceptance Criteria": "Data loader validates 22 columns, 0 missing cells; documents class prevalence (~9.3% positive); README lists all 21 predictor definitions from Table 2 (Thesis p.45–47).",
        "Status": "ToDo",
        "Research Trace": "Thesis Table 3–4; Notebooks read_csv BRFSS2015",
    },
    {
        "ID": "AC-011",
        "Repository": REPO_ML,
        "Epic": "E02 BRFSS Data Pipeline",
        "Functionality": "BRFSS 2024 LLCP extension path",
        "User Story": "As the ML engineer, I want an optional ingest path for LLCP2024 ASC/CSV CDC files with _MICHD outcome mapping, so that the journal methodology (2024 BRFSS, survey weights) can be reproduced.",
        "Role": "ML Engineer",
        "Type": "New",
        "Priority": "Should",
        "Phase": "Phase 3",
        "Source": "Journal §IV-A–C; ResearchPaper Data/LLCP2024",
        "Acceptance Criteria": "Parser maps CVDCRHD4+CVDINFR4→_MICHD binary target; applies _LLCPWT in evaluation scripts; documented divergence from 2015 Kaggle-cleaned set.",
        "Status": "ToDo",
        "Research Trace": "Journal Methodology IV-A–E",
    },
    {
        "ID": "AC-012",
        "Repository": REPO_ML,
        "Epic": "E02 BRFSS Data Pipeline",
        "Functionality": "Preprocessing & leakage-safe pipeline",
        "User Story": "As the ML engineer, I want sklearn Pipeline with median/mode imputation, StandardScaler/MinMaxScaler, one-hot encoding applied only on train fold, so that review concerns on data leakage are addressed.",
        "Role": "ML Engineer",
        "Type": "Existing",
        "Priority": "Must",
        "Phase": "Phase 1 — ML First",
        "Source": "Thesis §4.3–4.4; Journal §IV-E; Review Comments",
        "Acceptance Criteria": "Stratified 70/30 split (random_state=42); preprocessing fit on train only; unit tests assert no test statistics in fit; documents BMI/MentHlth/PhysHlth scaling rationale.",
        "Status": "ToDo",
        "Research Trace": "Thesis §4.3; Review (stratification, leakage)",
    },
    {
        "ID": "AC-013",
        "Repository": REPO_ML,
        "Epic": "E02 BRFSS Data Pipeline",
        "Functionality": "Exploratory data analysis artifacts",
        "User Story": "As a researcher, I want automated EDA reports (correlation heatmap, bivariate plots, class balance) matching thesis Chapter 4 visuals, so that the website documents dataset characteristics.",
        "Role": "Researcher",
        "Type": "Existing",
        "Priority": "Should",
        "Phase": "Phase 2",
        "Source": "Thesis §4.2.6–4.3.3; Notebooks pairplot/boxplot",
        "Acceptance Criteria": "Script emits PNG/SVG for top ±5 correlated features; publishes to /docs/assets/eda/; includes feature dictionary table.",
        "Status": "ToDo",
        "Research Trace": "Thesis Figures 8, 11; Notebook EDA cells",
    },
    # --- E03 ML models (implement ALL first — user requirement) ---
    {
        "ID": "AC-020",
        "Repository": REPO_ML,
        "Epic": "E03 Classifier Implementation",
        "Functionality": "Support Vector Machine (SVC)",
        "User Story": "As the ML engineer, I want a tuned SVC model with GridSearchCV hyperparameters, so that thesis Experiment 1 (~88.8% test accuracy) is reproducible.",
        "Role": "ML Engineer",
        "Type": "Existing",
        "Priority": "Must",
        "Phase": "Phase 1 — ML First",
        "Source": "CVD_SVM_.ipynb; Thesis §5.2.1",
        "Acceptance Criteria": "Train/eval script outputs accuracy, precision, recall, F1, ROC-AUC, confusion matrix; serialized model artifact; metrics within ±1% of thesis baseline or documented delta.",
        "Status": "ToDo",
        "Research Trace": "Thesis Exp 1; SVM notebook",
    },
    {
        "ID": "AC-021",
        "Repository": REPO_ML,
        "Epic": "E03 Classifier Implementation",
        "Functionality": "Decision Tree classifier",
        "User Story": "As the ML engineer, I want DecisionTreeClassifier with documented hyperparameters (Table 9), so that Experiment 2 results are reproducible.",
        "Role": "ML Engineer",
        "Type": "Existing",
        "Priority": "Must",
        "Phase": "Phase 1 — ML First",
        "Source": "CVD_DecisionTrees.ipynb; Thesis §5.2.2",
        "Acceptance Criteria": "Hyperparameter table in repo; AUC curve export; overfit gap (train vs test) reported explicitly.",
        "Status": "ToDo",
        "Research Trace": "Thesis Exp 2; Table 17 AUC",
    },
    {
        "ID": "AC-022",
        "Repository": REPO_ML,
        "Epic": "E03 Classifier Implementation",
        "Functionality": "k-Nearest Neighbors (KNN)",
        "User Story": "As the ML engineer, I want KNeighborsClassifier with NCA/PCA optional dimension reduction, so that Experiment 3 matches notebook workflow.",
        "Role": "ML Engineer",
        "Type": "Existing",
        "Priority": "Must",
        "Phase": "Phase 1 — ML First",
        "Source": "CVD_knn.ipynb; Thesis §5.2.3",
        "Acceptance Criteria": "GridSearch on k; metrics bundle exported; inference latency benchmark logged (review: computational cost).",
        "Status": "ToDo",
        "Research Trace": "Thesis Exp 3",
    },
    {
        "ID": "AC-023",
        "Repository": REPO_ML,
        "Epic": "E03 Classifier Implementation",
        "Functionality": "Neural Network (MLP)",
        "User Story": "As the ML engineer, I want an MLPClassifier (and optional Keras sequential baseline), so that NN experiment aligns with thesis §4.7.3 and excluded DL note §5.2.8.",
        "Role": "ML Engineer",
        "Type": "Existing",
        "Priority": "Must",
        "Phase": "Phase 1 — ML First",
        "Source": "CVD_NeuralNetworks.ipynb; Thesis §5.2.8",
        "Acceptance Criteria": "MLP trained with early stopping; DL sequential marked experimental if <80% accuracy; both reported in model registry.",
        "Status": "ToDo",
        "Research Trace": "Thesis §4.7.3, §5.2.8",
    },
    {
        "ID": "AC-024",
        "Repository": REPO_ML,
        "Epic": "E03 Classifier Implementation",
        "Functionality": "Random Forest",
        "User Story": "As the ML engineer, I want RandomForestClassifier with thesis hyperparameters (Table 11), so that Experiment 4 (~89% test accuracy) is reproduced.",
        "Role": "ML Engineer",
        "Type": "Existing",
        "Priority": "Must",
        "Phase": "Phase 1 — ML First",
        "Source": "CVD_RandomForest.ipynb; Thesis §5.2.4",
        "Acceptance Criteria": "Feature importances export; OOB score if enabled; metrics JSON committed to /models/metrics/rf.json.",
        "Status": "ToDo",
        "Research Trace": "Thesis Exp 4; Table 18 AUC",
    },
    {
        "ID": "AC-025",
        "Repository": REPO_ML,
        "Epic": "E03 Classifier Implementation",
        "Functionality": "XGBoost / Gradient Boosting",
        "User Story": "As the ML engineer, I want XGBoost Random Forest / gradient boosting experiment, so that thesis Experiment 5 (best test accuracy claim) is included in comparison.",
        "Role": "ML Engineer",
        "Type": "Existing",
        "Priority": "Must",
        "Phase": "Phase 1 — ML First",
        "Source": "Thesis §5.2.5",
        "Acceptance Criteria": "XGBoost model trained; comparison row in leaderboard; documents thesis selection vs journal ensemble selection difference.",
        "Status": "ToDo",
        "Research Trace": "Thesis Exp 5 — XGBoost RF best test acc",
    },
    {
        "ID": "AC-026",
        "Repository": REPO_ML,
        "Epic": "E03 Classifier Implementation",
        "Functionality": "Ensemble (Voting / Bagging / Boosting)",
        "User Story": "As the ML engineer, I want ensemble models (Majority Voting, Weighted Voting, Bagging per §4.7.6.1), so that journal's 89.1% ensemble result is the default production model.",
        "Role": "ML Engineer",
        "Type": "Existing",
        "Priority": "Must",
        "Phase": "Phase 1 — ML First",
        "Source": "CVD_Ensembles.ipynb; Thesis §4.7.6; Journal §V-A",
        "Acceptance Criteria": "VotingClassifier over top base estimators; 89.1% test accuracy reproduced or explained; ensemble weights persisted; default model flag in registry.",
        "Status": "ToDo",
        "Research Trace": "Thesis Table 12; Journal selects Ensemble",
    },
    {
        "ID": "AC-027",
        "Repository": REPO_ML,
        "Epic": "E03 Classifier Implementation",
        "Functionality": "ANFIS / fuzzy baseline",
        "User Story": "As the ML engineer, I want an ANFIS or scikit-fuzzy baseline, so that literature-covered fuzzy logic approach (Thesis §2.4.3) is represented alongside sklearn classifiers.",
        "Role": "ML Engineer",
        "Type": "Existing",
        "Priority": "Should",
        "Phase": "Phase 1 — ML First",
        "Source": "CVD_ANFIS .ipynb; Thesis §2.4.3",
        "Acceptance Criteria": "ANFIS notebook logic ported to pipeline script OR documented as research-only if MATLAB dependency; included in model comparison table.",
        "Status": "ToDo",
        "Research Trace": "ANFIS notebook; Thesis fuzzy/ANFIS review",
    },
    {
        "ID": "AC-028",
        "Repository": REPO_ML,
        "Epic": "E03 Classifier Implementation",
        "Functionality": "Logistic Regression & Naive Bayes baselines",
        "User Story": "As the ML engineer, I want calibrated LogisticRegression and GaussianNB baselines, so that thesis Experiments 6–7 and review-requested interpretable baselines exist.",
        "Role": "ML Engineer",
        "Type": "Existing",
        "Priority": "Must",
        "Phase": "Phase 1 — ML First",
        "Source": "Thesis §5.2.6–5.2.7; Review Comments",
        "Acceptance Criteria": "Logistic regression selected in thesis for explainability — expose as explainability-friendly baseline; Platt/isotonic calibration curve exported.",
        "Status": "ToDo",
        "Research Trace": "Thesis §5.3 — LR best for production criteria",
    },
    # --- E04 Evaluation ---
    {
        "ID": "AC-030",
        "Repository": REPO_ML,
        "Epic": "E04 Training, Evaluation & Model Selection",
        "Functionality": "Unified metrics & leaderboard",
        "User Story": "As the product owner, I want a single leaderboard comparing all models on accuracy, precision, recall, F1, ROC-AUC, train/test gap, so that thesis §5.3 selection criteria are transparent.",
        "Role": "Product Owner",
        "Type": "New",
        "Priority": "Must",
        "Phase": "Phase 1 — ML First",
        "Source": "Thesis §5.3; Journal Table 1; Review Comments",
        "Acceptance Criteria": "metrics.json + HTML report; emphasizes recall/AUC under imbalance; qualitative dimensions (explainability, complexity, cost) from Journal Table 1.",
        "Status": "ToDo",
        "Research Trace": "Thesis Figs 55–66; Journal Fig 1",
    },
    {
        "ID": "AC-031",
        "Repository": REPO_ML,
        "Epic": "E04 Training, Evaluation & Model Selection",
        "Functionality": "Stratified CV & hyperparameter tuning",
        "User Story": "As the ML engineer, I want stratified k-fold CV with nested hyperparameter search, so that peer-review gaps on validation protocol are closed.",
        "Role": "ML Engineer",
        "Type": "New",
        "Priority": "Must",
        "Phase": "Phase 1 — ML First",
        "Source": "Review Comments 01.txt",
        "Acceptance Criteria": "5-fold stratified CV; bootstrap 95% CI for AUROC; no test fold in tuning; documented in METHODOLOGY.md.",
        "Status": "ToDo",
        "Research Trace": "Review — nested CV, CI",
    },
    {
        "ID": "AC-032",
        "Repository": REPO_ML,
        "Epic": "E04 Training, Evaluation & Model Selection",
        "Functionality": "Calibration & decision-curve analysis",
        "User Story": "As a clinician reviewer, I want reliability curves, Brier score, and decision-curve analysis for the production model, so that screening-tool claims are clinically interpretable.",
        "Role": "Clinician",
        "Type": "New",
        "Priority": "Must",
        "Phase": "Phase 2",
        "Source": "Review Comments 01.txt; Journal §V-C",
        "Acceptance Criteria": "Calibration plot PNG; DCA net benefit chart; stated operating threshold with confusion matrix at that threshold.",
        "Status": "ToDo",
        "Research Trace": "Review — calibration, DCA missing",
    },
    {
        "ID": "AC-033",
        "Repository": REPO_ML,
        "Epic": "E04 Training, Evaluation & Model Selection",
        "Functionality": "Clinical risk score benchmarks",
        "User Story": "As a researcher, I want comparison vs Framingham / ASCVD PCE baseline on overlapping features, so that incremental ML value is quantified.",
        "Role": "Researcher",
        "Type": "New",
        "Priority": "Should",
        "Phase": "Phase 3",
        "Source": "Review Comments 01.txt; Journal Limitations",
        "Acceptance Criteria": "Benchmark notebook; table of AUROC for clinical scores vs ensemble; limitations documented if feature parity incomplete.",
        "Status": "ToDo",
        "Research Trace": "Review — Framingham/ASCVD baseline",
    },
    {
        "ID": "AC-034",
        "Repository": REPO_ML,
        "Epic": "E04 Training, Evaluation & Model Selection",
        "Functionality": "Fairness / subgroup evaluation",
        "User Story": "As an ethics reviewer, I want AUROC and calibration sliced by sex, age band, race/ethnicity, income, education, so that BRFSS equity concerns are measurable.",
        "Role": "Ethics Reviewer",
        "Type": "New",
        "Priority": "Should",
        "Phase": "Phase 3",
        "Source": "Review Comments; Journal §VI-A",
        "Acceptance Criteria": "Subgroup metrics table; disparity flags when ΔAUROC > 0.05; documented mitigation plan.",
        "Status": "ToDo",
        "Research Trace": "Review — subgroup fairness",
    },
    {
        "ID": "AC-035",
        "Repository": REPO_ML,
        "Epic": "E04 Training, Evaluation & Model Selection",
        "Functionality": "Model export for web inference",
        "User Story": "As the web developer, I want serialized models (joblib/onnx) plus preprocessing manifest, so that GitHub Pages can run inference without a Python server.",
        "Role": "Web Developer",
        "Type": "New",
        "Priority": "Must",
        "Phase": "Phase 1 — ML First",
        "Source": "Product owner — GitHub Pages deployment",
        "Acceptance Criteria": "All production models exported to /models/artifacts/; inference schema JSON; Pyodide or ONNX.js proof-of-concept documented with latency budget <3s.",
        "Status": "ToDo",
        "Research Trace": "Thesis §4.9 MLOps pipeline proposal",
    },
    # --- E05 Explainability ---
    {
        "ID": "AC-040",
        "Repository": REPO_ML,
        "Epic": "E05 Explainable AI (XAI)",
        "Functionality": "ELI5 global explanations",
        "User Story": "As a healthcare professional, I want global feature weights/importance from ELI5 for the selected model, so that thesis §5.4.1 global explainability is available in the product.",
        "Role": "Clinician",
        "Type": "Existing",
        "Priority": "Must",
        "Phase": "Phase 1 — ML First",
        "Source": "Thesis §5.4.1; Journal §IV-G",
        "Acceptance Criteria": "Top features (Age, BMI, Smoker, HighBP, Diabetes) ranked; HTML/text export; matches thesis Figure 67 ordering qualitatively.",
        "Status": "ToDo",
        "Research Trace": "Thesis Fig 67; Journal Fig 2",
    },
    {
        "ID": "AC-041",
        "Repository": REPO_ML,
        "Epic": "E05 Explainable AI (XAI)",
        "Functionality": "ELI5 local instance explanations",
        "User Story": "As a patient, I want a per-prediction ELI5 breakdown showing contributing features and probability, so that thesis §5.4.2 local examples are reproducible in the UI.",
        "Role": "Patient",
        "Type": "Existing",
        "Priority": "Must",
        "Phase": "Phase 2",
        "Source": "Thesis §5.4.2; Journal §V-B",
        "Acceptance Criteria": "API/function show_prediction() returns top ± contributors; three canonical test cases from thesis §5.4.2.1–3 included as demos.",
        "Status": "ToDo",
        "Research Trace": "Thesis Figs 68–73",
    },
    {
        "ID": "AC-042",
        "Repository": REPO_ML,
        "Epic": "E05 Explainable AI (XAI)",
        "Functionality": "SHAP complement (Phase 3)",
        "User Story": "As a researcher, I want optional SHAP summary and force plots for tree ensembles, so that review recommendation to complement ELI5 is addressed.",
        "Role": "Researcher",
        "Type": "New",
        "Priority": "Could",
        "Phase": "Phase 3",
        "Source": "Review Comments; Journal §VI-B",
        "Acceptance Criteria": "SHAP plots for RF/Ensemble; side-by-side ELI5 vs SHAP page; faithfulness note in docs.",
        "Status": "ToDo",
        "Research Trace": "Review — SHAP over ELI5 for trees",
    },
    # --- E06 Website ---
    {
        "ID": "AC-050",
        "Repository": REPO_WEB,
        "Epic": "E06 Web Application Shell",
        "Functionality": "Marketing & research landing page",
        "User Story": "As a visitor, I want a landing page summarizing the AI Cardiologist research (problem, dataset, 89.1% ensemble result, XAI approach), so that the site serves as the public face of the thesis.",
        "Role": "Visitor",
        "Type": "New",
        "Priority": "Must",
        "Phase": "MVP",
        "Source": "Journal Abstract; Thesis Abstract",
        "Acceptance Criteria": "Hero, CTA to Predict & Research sections; mobile-responsive; links to GitHub repo and PDF downloads.",
        "Status": "ToDo",
        "Research Trace": "Thesis Abstract; Journal §I",
    },
    {
        "ID": "AC-051",
        "Repository": REPO_WEB,
        "Epic": "E06 Web Application Shell",
        "Functionality": "Site IA & navigation",
        "User Story": "As a visitor, I want clear navigation across Predict, Models, Research, References, and About, so that ML capabilities and documentation are discoverable.",
        "Role": "Visitor",
        "Type": "New",
        "Priority": "Must",
        "Phase": "MVP",
        "Source": "FrenchCoachPro DOC hub pattern",
        "Acceptance Criteria": "Top nav + footer; breadcrumb on inner pages; matches README information architecture.",
        "Status": "ToDo",
        "Research Trace": "FrenchCoachPro.Doc ECOSYSTEM.md pattern",
    },
    {
        "ID": "AC-052",
        "Repository": REPO_WEB,
        "Epic": "E06 Web Application Shell",
        "Functionality": "Accessible UI (WCAG 2.1 AA)",
        "User Story": "As a user with assistive technology, I want forms and results to meet WCAG 2.1 AA, so that healthcare information is equitably accessible.",
        "Role": "Patient",
        "Type": "New",
        "Priority": "Should",
        "Phase": "Phase 2",
        "Source": "FUTURE-AI guideline; Ethics refs",
        "Acceptance Criteria": "axe-core CI check; keyboard navigable forms; color contrast ≥4.5:1 on risk indicators.",
        "Status": "ToDo",
        "Research Trace": "REF_7 FUTURE-AI",
    },
    # --- E07 User testing / prediction ---
    {
        "ID": "AC-060",
        "Repository": REPO_WEB,
        "Epic": "E07 Prediction & User Testing",
        "Functionality": "Interactive BRFSS feature intake form",
        "User Story": "As a user, I want to enter all 21 BRFSS predictors via a guided form with tooltips from the thesis feature dictionary, so that I can test the model with my own survey answers.",
        "Role": "User / Patient",
        "Type": "New",
        "Priority": "Must",
        "Phase": "Phase 2",
        "Source": "Thesis Table 2; User requirement",
        "Acceptance Criteria": "Fields: HighBP, HighChol, CholCheck, BMI, Smoker, Stroke, Diabetes, PhysActivity, Fruits, Veggies, HvyAlcoholConsump, AnyHealthcare, NoDocbcCost, GenHlth, MentHlth, PhysHlth, DiffWalk, Sex, Age, Education, Income; validation rules; sensible defaults.",
        "Status": "ToDo",
        "Research Trace": "Thesis §4.2.2 feature list",
    },
    {
        "ID": "AC-061",
        "Repository": REPO_WEB,
        "Epic": "E07 Prediction & User Testing",
        "Functionality": "Model selector & ensemble default",
        "User Story": "As a user, I want to choose which trained classifier runs my prediction (default Ensemble), so that I can compare algorithm behavior interactively.",
        "Role": "User",
        "Type": "New",
        "Priority": "Must",
        "Phase": "Phase 2",
        "Source": "Thesis §5.3; User requirement — all ML methods",
        "Acceptance Criteria": "Dropdown lists all AC-020–AC-028 models; shows latency; Ensemble pre-selected; disabled models show reason.",
        "Status": "ToDo",
        "Research Trace": "All CVD_*.ipynb models",
    },
    {
        "ID": "AC-062",
        "Repository": REPO_WEB,
        "Epic": "E07 Prediction & User Testing",
        "Functionality": "Prediction result & risk visualization",
        "User Story": "As a user, I want CVD risk probability, binary class, and plain-language interpretation, so that I understand the output without clinical training.",
        "Role": "User",
        "Type": "New",
        "Priority": "Must",
        "Phase": "Phase 2",
        "Source": "Journal §V-B; Thesis explainability goals",
        "Acceptance Criteria": "Shows P(CVD), threshold, recall-oriented warning if high FN cost; red/amber/green with text alternative; links to ELI5 local explanation.",
        "Status": "ToDo",
        "Research Trace": "Journal Results §V",
    },
    {
        "ID": "AC-063",
        "Repository": REPO_WEB,
        "Epic": "E07 Prediction & User Testing",
        "Functionality": "Batch CSV upload & test harness",
        "User Story": "As a researcher, I want to upload a CSV (BRFSS schema) and receive bulk predictions + metrics, so that I can test with sample or holdout data.",
        "Role": "Researcher",
        "Type": "New",
        "Priority": "Must",
        "Phase": "Phase 2",
        "Source": "User requirement — test with data",
        "Acceptance Criteria": "Client-side parse ≤10MB; outputs downloadable results CSV; if labels present, shows confusion matrix & AUROC; sample dataset link to BRFSS2015 subset.",
        "Status": "ToDo",
        "Research Trace": "ResearchPaper Data/; Notebooks",
    },
    {
        "ID": "AC-064",
        "Repository": REPO_WEB,
        "Epic": "E07 Prediction & User Testing",
        "Functionality": "Model comparison dashboard",
        "User Story": "As a researcher, I want interactive charts mirroring thesis §5.3 bar charts (accuracy, F1, AUC across models), so that research findings are explorable on the site.",
        "Role": "Researcher",
        "Type": "New",
        "Priority": "Should",
        "Phase": "Phase 2",
        "Source": "Thesis Figs 55–66; Journal Fig 1",
        "Acceptance Criteria": "Charts from metrics.json; toggle train vs test; export PNG.",
        "Status": "ToDo",
        "Research Trace": "Thesis §5.3.2 ranking charts",
    },
    {
        "ID": "AC-065",
        "Repository": REPO_WEB,
        "Epic": "E07 Prediction & User Testing",
        "Functionality": "Demo scenarios (thesis case studies)",
        "User Story": "As a visitor, I want one-click load of the three thesis local-explanation examples, so that I can see explainability without entering data manually.",
        "Role": "Visitor",
        "Type": "New",
        "Priority": "Should",
        "Phase": "Phase 2",
        "Source": "Thesis §5.4.2",
        "Acceptance Criteria": "Three preset profiles; expected probability ~0.628, 0.832, 0.693 documented.",
        "Status": "ToDo",
        "Research Trace": "Thesis §5.4.2.1–3",
    },
    # --- E08 Research documentation hub ---
    {
        "ID": "AC-070",
        "Repository": REPO_DOCS,
        "Epic": "E08 Research Documentation Hub",
        "Functionality": "Thesis & journal document viewer",
        "User Story": "As a researcher, I want embedded/downloadable MS Navendu Brajesh V1.2 and AI_Cardiologist_for_journals PDFs, so that full research is available alongside the demo.",
        "Role": "Researcher",
        "Type": "Existing",
        "Priority": "Must",
        "Phase": "MVP",
        "Source": "AI Cardiologist.Doc PDFs",
        "Acceptance Criteria": "PDF links (not committed if large — GitHub release assets); metadata (author, date, abstract); version V1.2 labeled.",
        "Status": "ToDo",
        "Research Trace": "MS Navendu Brajesh V1.2.pdf",
    },
    {
        "ID": "AC-071",
        "Repository": REPO_DOCS,
        "Epic": "E08 Research Documentation Hub",
        "Functionality": "Methodology documentation suite",
        "User Story": "As a developer, I want README, ARCHITECTURE, DATA_MODEL, METHODOLOGY, and EVALUATION markdown docs, so that the repo matches FrenchCoachPro.Doc engineering standards.",
        "Role": "Developer",
        "Type": "New",
        "Priority": "Must",
        "Phase": "MVP",
        "Source": "FrenchCoachPro.Doc structure",
        "Acceptance Criteria": "docs/ folder with listed files; DATA_MODEL enumerates 21 features; METHODOLOGY cites label HeartDiseaseorAttack / _MICHD.",
        "Status": "ToDo",
        "Research Trace": "FrenchCoachPro ReadMe.txt suite list",
    },
    {
        "ID": "AC-072",
        "Repository": REPO_DOCS,
        "Epic": "E08 Research Documentation Hub",
        "Functionality": "Literature review & gap analysis page",
        "User Story": "As a researcher, I want a structured lit-review page covering SVM, DT, NN, KNN, RF, Ensembles, ANFIS, XAI, ethics, so that thesis Chapter 2 is navigable on the web.",
        "Role": "Researcher",
        "Type": "Existing",
        "Priority": "Should",
        "Phase": "Phase 2",
        "Source": "Thesis Ch.2; ReferencePaper/",
        "Acceptance Criteria": "Sections map to thesis headings; arXiv/DOI links from RefrenceURLList.txt; 12 reference PDFs in ReferencePaper indexed.",
        "Status": "ToDo",
        "Research Trace": "Thesis Ch.2; ReferencePaper/*.pdf",
    },
    {
        "ID": "AC-073",
        "Repository": REPO_DOCS,
        "Epic": "E08 Research Documentation Hub",
        "Functionality": "Peer review & remediation log",
        "User Story": "As the author, I want Review Comments 01.txt tracked as remediation stories (AC-031, AC-032, etc.), so that journal resubmission gaps are visible.",
        "Role": "Author",
        "Type": "New",
        "Priority": "Should",
        "Phase": "Phase 3",
        "Source": "Review/Review Comments 01.txt",
        "Acceptance Criteria": "Review page lists weakness → story ID → status; links to evidence artifacts.",
        "Status": "ToDo",
        "Research Trace": "Review Comments 01.txt",
    },
    {
        "ID": "AC-074",
        "Repository": REPO_DOCS,
        "Epic": "E08 Research Documentation Hub",
        "Functionality": "MLOps pipeline documentation",
        "User Story": "As an ML engineer, I want thesis §4.9 MLOps tool mapping (Table 14–15) translated into a deployable CI outline, so that Objective 4 (MLOps proposal) is realized.",
        "Role": "ML Engineer",
        "Type": "Existing",
        "Priority": "Should",
        "Phase": "Phase 3",
        "Source": "Thesis §4.9; Fig 19 Pipeline",
        "Acceptance Criteria": "MLOPS.md describes train → validate → export → deploy; GitHub Actions workflow stub.",
        "Status": "ToDo",
        "Research Trace": "Thesis §4.9, §6.1.4",
    },
    # --- E09 Deployment ---
    {
        "ID": "AC-080",
        "Repository": REPO_WEB,
        "Epic": "E09 GitHub Pages Deployment",
        "Functionality": "GitHub Pages static site pipeline",
        "User Story": "As the maintainer, I want GitHub Actions to build and deploy the site to GitHub Pages on main, so that the demo is publicly accessible.",
        "Role": "Maintainer",
        "Type": "New",
        "Priority": "Must",
        "Phase": "MVP",
        "Source": "User — GitHub Site deployment",
        "Acceptance Criteria": "workflow deploys Vite/React or similar static build; custom domain doc; base path configured; Lighthouse perf ≥80.",
        "Status": "ToDo",
        "Research Trace": "Product owner deployment intent",
    },
    {
        "ID": "AC-081",
        "Repository": REPO_ML,
        "Epic": "E09 GitHub Pages Deployment",
        "Functionality": "CI train & export workflow",
        "User Story": "As the ML engineer, I want CI to retrain or validate models on demand and commit metrics artifacts, so that the website stays synchronized with code.",
        "Role": "ML Engineer",
        "Type": "New",
        "Priority": "Should",
        "Phase": "Phase 3",
        "Source": "Thesis MLOps; AC-074",
        "Acceptance Criteria": "workflow_dispatch job; caches BRFSS CSV; fails if metrics regress >2% AUROC vs baseline.",
        "Status": "ToDo",
        "Research Trace": "Thesis §4.9 pipeline",
    },
    {
        "ID": "AC-082",
        "Repository": REPO_WEB,
        "Epic": "E09 GitHub Pages Deployment",
        "Functionality": "Client-side inference strategy",
        "User Story": "As the architect, I want documented client-side inference (Pyodide/sklearn WASM or ONNX.js) because GitHub Pages is static, so that predictions work without a paid backend.",
        "Role": "Architect",
        "Type": "New",
        "Priority": "Must",
        "Phase": "Phase 1 — ML First",
        "Source": "GitHub Pages constraint",
        "Acceptance Criteria": "SPIKE doc compares Pyodide vs ONNX.js vs optional Render API; decision recorded in ARCHITECTURE.md; POC demo in /predict.",
        "Status": "ToDo",
        "Research Trace": "Deployment architecture decision",
    },
    # --- E10 Hypotheses & validation ---
    {
        "ID": "AC-090",
        "Repository": REPO_DOCS,
        "Epic": "E10 Research Traceability",
        "Functionality": "Hypothesis H1–H4 traceability matrix",
        "User Story": "As a reviewer, I want each hypothesis linked to UI features and metric artifacts, so that journal §II-C claims are verifiable.",
        "Role": "Reviewer",
        "Type": "New",
        "Priority": "Should",
        "Phase": "Phase 2",
        "Source": "Journal §II-C",
        "Acceptance Criteria": "Matrix: H1→AC-030/062; H2→AC-040/064; H3→AC-041; H4→disclaimer+theoretical note until health econ study.",
        "Status": "ToDo",
        "Research Trace": "Journal Hypotheses H1–H4",
    },
    {
        "ID": "AC-091",
        "Repository": REPO_DOCS,
        "Epic": "E10 Research Traceability",
        "Functionality": "Healthcare professional validation module (future)",
        "User Story": "As a clinician, I want a structured feedback form on explanation quality (Likert), so that thesis Objective 3 validation can be collected prospectively.",
        "Role": "Clinician",
        "Type": "New",
        "Priority": "Could",
        "Phase": "Enterprise",
        "Source": "Thesis §1.6.4; Review Questions Q7",
        "Acceptance Criteria": "Optional survey instrument; stores anonymized responses; Max Healthcare attribution in acknowledgements.",
        "Status": "ToDo",
        "Research Trace": "Thesis Acknowledgements — Max Healthcare",
    },
]


def write_read_me(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Read Me", 0)
    lines = [
        ("AI Cardiologist — Product Requirements", ""),
        ("Document", "AI Cardiologist Requirement.xlsx"),
        ("Generated", str(date.today())),
        ("Author / Transition", "Navendu Brajesh — Research → Product (Consulting backlog)"),
        ("Primary Research", "MS Navendu Brajesh V1.2.pdf (LJMU Final Thesis, Aug 2022)"),
        ("Journal Draft", "AI_Cardiologist_for_journals.pdf"),
        ("", ""),
        ("PRODUCT VISION", ""),
        (
            "Goal",
            "Deliver a working public website (GitHub Pages) that implements ALL thesis ML classifiers "
            "first, then enables users to test predictions with their own or CSV data, with full research "
            "documentation embedded.",
        ),
        (
            "Clinical framing",
            "Decision-support / population screening only — NOT a diagnostic device. Human cardiologist in the loop.",
        ),
        ("", ""),
        ("REPOSITORIES (TARGET STATE)", ""),
        (REPO_ML, "Python training pipeline, notebooks port, model export, metrics, XAI generation."),
        (REPO_WEB, "Static SPA (Vite+React recommended) — predict, batch test, dashboards, deploy to GitHub Pages."),
        (REPO_DOCS, "Research hub — thesis, journal, references, methodology, review remediation."),
        ("", ""),
        ("BUILD ORDER (OWNER PRIORITY)", ""),
        ("Phase 1 — ML First", "AC-010–AC-035, AC-040, AC-082: data pipeline, all 8+ classifiers, ensemble, ELI5 global, export artifacts."),
        ("Phase 2 — User Testing", "AC-060–AC-065, AC-041–AC-042, AC-032: web forms, batch CSV, results, local XAI, calibration views."),
        ("MVP (parallel)", "AC-001–AC-002, AC-050–AC-051, AC-070–AC-071, AC-080: landing, governance, docs shell, GitHub Pages."),
        ("Phase 3", "AC-011, AC-033–AC-034, AC-042, AC-072–AC-074, AC-081: 2024 BRFSS, fairness, SHAP, MLOps CI, clinical benchmarks."),
        ("Enterprise", "AC-091: prospective clinician validation studies."),
        ("", ""),
        ("KEY RESEARCH FACTS (BASELINE)", ""),
        ("Dataset (thesis/notebooks)", "heart_disease_health_indicators_BRFSS2015.csv — 253,680 rows, 21 features, target HeartDiseaseorAttack"),
        ("Dataset (journal)", "BRFSS 2024 LLCP — ~500k rows, target _MICHD (CHD/MI composite), survey weight _LLCPWT"),
        ("Best thesis test accuracy", "XGBoost Random Forest (~91% test); Logistic Regression selected for production explainability"),
        ("Journal headline result", "Ensemble 89.1% test accuracy with ELI5 explanations"),
        ("Top predictors", "Age, BMI, Smoker, Hypertension (HighBP), Diabetes"),
        ("XAI approach", "ELI5 global weights + local instance explanations (SHAP recommended in review)"),
        ("", ""),
        ("PEER REVIEW GAPS → REQUIREMENTS", ""),
        ("Label definition", "AC-010/071 — explicit HeartDiseaseorAttack / _MICHD documentation"),
        ("Stratified CV / CI", "AC-031"),
        ("Calibration / DCA", "AC-032"),
        ("Clinical score baseline", "AC-033"),
        ("Subgroup fairness", "AC-034"),
        ("ELI5 faithfulness", "AC-041, AC-042"),
        ("External validation", "Deferred — AC-011 partial; NHANES/multi-year BRFSS noted in limitations"),
        ("", ""),
        ("TECH STACK (RECOMMENDED)", ""),
        ("ML", "Python 3.11+, scikit-learn, xgboost, imbalanced-learn, eli5, shap, pandas"),
        ("Web", "Vite + React + TypeScript + Tailwind; Pyodide or ONNX.js for client inference"),
        ("Deploy", "GitHub Actions → GitHub Pages; model artifacts in repo or Git LFS"),
        ("Docs", "Markdown suite mirroring FrenchCoachPro.Doc (README, ARCHITECTURE, DATA_MODEL, etc.)"),
    ]
    for i, (a, b) in enumerate(lines, 1):
        ws.cell(i, 1, a)
        ws.cell(i, 2, b)
        if a in {
            "PRODUCT VISION", "REPOSITORIES (TARGET STATE)", "BUILD ORDER (OWNER PRIORITY)",
            "KEY RESEARCH FACTS (BASELINE)", "PEER REVIEW GAPS → REQUIREMENTS", "TECH STACK (RECOMMENDED)",
        }:
            ws.cell(i, 1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 100


def write_architecture(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Architecture")
    hdr = ["Layer", "Component", "Technology", "Rationale"]
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    rows = [
        ("Data", "BRFSS 2015 CSV", "CDC/Kaggle cleaned", "Primary thesis replication set"),
        ("Data", "BRFSS 2024 LLCP", "CDC ASC/CSV/XPT", "Journal methodology alignment"),
        ("ML Train", "Training scripts", "Python + sklearn + XGBoost", "Port from CVD_*.ipynb"),
        ("ML Train", "Experiment tracking", "metrics.json + MLflow optional", "Reproducible leaderboard"),
        ("ML Serve", "Model artifacts", "joblib → ONNX or Pyodide", "Static hosting constraint"),
        ("XAI", "ELI5 / SHAP", "eli5, shap", "Thesis + review requirements"),
        ("Web", "SPA", "Vite + React", "GitHub Pages compatible"),
        ("Web", "Inference", "Client-side WASM", "No backend required for MVP"),
        ("Docs", "Research hub", "Markdown + PDF assets", "Thesis/journal/reference library"),
        ("CI/CD", "Deploy", "GitHub Actions", "Pages deploy + optional retrain job"),
    ]
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            ws.cell(i, j, val)
    for col, w in enumerate([12, 28, 28, 48], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def write_data_model(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Data Model")
    hdr = ["Feature", "Type", "Description", "Values / Range", "Normalize"]
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    features = [
        ("HeartDiseaseorAttack", "Target (int)", "CHD or MI ever reported", "0=No, 1=Yes", "N/A"),
        ("HighBP", "int", "Told has high blood pressure", "0/1", "No"),
        ("HighChol", "int", "Told has high cholesterol", "0/1", "No"),
        ("CholCheck", "int", "Cholesterol checked in past 5 years", "0/1", "No"),
        ("BMI", "float/int", "Body Mass Index", "12–98", "Yes"),
        ("Smoker", "int", "Smoked 100+ cigarettes lifetime", "0/1", "No"),
        ("Stroke", "int", "Ever told had stroke", "0/1", "No"),
        ("Diabetes", "int", "Diabetes status", "0/1/2 (incl. gestational)", "No"),
        ("PhysActivity", "int", "Physical activity past 30 days", "0/1", "No"),
        ("Fruits", "int", "Fruit ≥1 time/day", "0/1", "No"),
        ("Veggies", "int", "Vegetables ≥1 time/day", "0/1", "No"),
        ("HvyAlcoholConsump", "int", "Heavy alcohol consumption", "0/1", "No"),
        ("AnyHealthcare", "int", "Has health care coverage", "0/1", "No"),
        ("NoDocbcCost", "int", "Could not see doctor due to cost", "0/1", "No"),
        ("GenHlth", "int", "General health status", "1–5 (Excellent→Poor)", "No"),
        ("MentHlth", "int", "Mental health not good days (30d)", "0–30", "Yes"),
        ("PhysHlth", "int", "Physical health not good days (30d)", "0–30", "Yes"),
        ("DiffWalk", "int", "Serious difficulty walking/climbing", "0/1", "No"),
        ("Sex", "int", "Sex", "0=Female, 1=Male", "No"),
        ("Age", "int", "Age category (14 levels)", "1–14 (18–80+)", "No"),
        ("Education", "int", "Education level", "1–6", "No"),
        ("Income", "int", "Annual household income band", "1–8", "No"),
    ]
    for i, row in enumerate(features, 2):
        for j, val in enumerate(row, 1):
            ws.cell(i, j, val)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 12


def write_sheet(ws, rows: list[dict], header: list[str]) -> None:
    for c, h in enumerate(header, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ri, row in enumerate(rows, 2):
        for ci, key in enumerate(header, 1):
            val = row.get(key, "")
            cell = ws.cell(ri, ci, val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if key == "Status":
                if val == "Done":
                    cell.fill = DONE_FILL
                elif row.get("Phase") == "Enterprise":
                    cell.fill = DEFER_FILL
                else:
                    cell.fill = TODO_FILL

    widths = {
        1: 10, 2: 34, 3: 34, 4: 28, 5: 52, 6: 16, 7: 12, 8: 10, 9: 18,
        10: 28, 11: 52, 12: 10, 13: 32,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(rows) + 1}"


def write_epic_summary(wb: openpyxl.Workbook, backlog: list[dict]) -> None:
    ws = wb.create_sheet("Epic Summary")
    hdr = ["Epic", "Total", "Done", "ToDo", "Must ToDo", "% Complete"]
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    stats: dict[str, dict] = defaultdict(lambda: {"total": 0, "done": 0, "todo": 0, "must_todo": 0})
    for row in backlog:
        epic = row["Epic"]
        stats[epic]["total"] += 1
        if row["Status"] == "Done":
            stats[epic]["done"] += 1
        else:
            stats[epic]["todo"] += 1
            if row["Priority"] == "Must":
                stats[epic]["must_todo"] += 1

    ri = 2
    total = {"total": 0, "done": 0, "todo": 0, "must_todo": 0}
    for epic in sorted(stats.keys()):
        s = stats[epic]
        ws.cell(ri, 1, epic)
        ws.cell(ri, 2, s["total"])
        ws.cell(ri, 3, s["done"])
        ws.cell(ri, 4, s["todo"])
        ws.cell(ri, 5, s["must_todo"])
        pct = round(100 * s["done"] / s["total"], 1) if s["total"] else 0
        ws.cell(ri, 6, pct)
        for k in total:
            total[k] += s[k]
        ri += 1
    ws.cell(ri, 1, "TOTAL")
    ws.cell(ri, 2, total["total"])
    ws.cell(ri, 3, total["done"])
    ws.cell(ri, 4, total["todo"])
    ws.cell(ri, 5, total["must_todo"])
    ws.cell(ri, 6, 0.0)
    ws.cell(ri, 1).font = Font(bold=True)


def write_roadmap(wb: openpyxl.Workbook, backlog: list[dict]) -> None:
    ws = wb.create_sheet("Roadmap")
    hdr = ["Phase", "Repository", "ID", "Functionality", "Priority", "Status"]
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    phase_order = [
        "Phase 1 — ML First",
        "MVP",
        "Phase 2",
        "Phase 3",
        "Enterprise",
    ]
    ri = 2
    for phase in phase_order:
        items = [b for b in backlog if b.get("Phase") == phase]
        for row in sorted(items, key=lambda x: (x["Priority"], x["ID"])):
            ws.cell(ri, 1, phase)
            ws.cell(ri, 2, row["Repository"])
            ws.cell(ri, 3, row["ID"])
            ws.cell(ri, 4, row["Functionality"])
            ws.cell(ri, 5, row["Priority"])
            ws.cell(ri, 6, row["Status"])
            ri += 1
    for col, w in enumerate([22, 34, 10, 36, 10, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def main() -> None:
    backlog = BACKLOG
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_read_me(wb)
    write_architecture(wb)
    write_data_model(wb)
    ws = wb.create_sheet("Backlog")
    write_sheet(ws, backlog, HEADER)
    write_epic_summary(wb, backlog)
    write_roadmap(wb, backlog)
    wb.save(OUT_XLSX)

    must = sum(1 for b in backlog if b["Priority"] == "Must")
    p1 = sum(1 for b in backlog if b.get("Phase") == "Phase 1 — ML First")
    print(f"Created {OUT_XLSX}")
    print(f"Stories: {len(backlog)} | Must: {must} | Phase 1 (ML First): {p1}")


if __name__ == "__main__":
    main()
