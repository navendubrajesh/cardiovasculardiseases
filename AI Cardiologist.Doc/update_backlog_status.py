"""Mark implemented backlog items as Done in AI Cardiologist Requirement.xlsx."""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "AI Cardiologist.Doc" / "AI Cardiologist Requirement.xlsx"

DONE = {
    "AC-001", "AC-002", "AC-010", "AC-011", "AC-012", "AC-013",
    "AC-020", "AC-021", "AC-022", "AC-023", "AC-024", "AC-025", "AC-026", "AC-027", "AC-028",
    "AC-030", "AC-031", "AC-035", "AC-040", "AC-041",
    "AC-050", "AC-051", "AC-060", "AC-061", "AC-062", "AC-063", "AC-064", "AC-065",
    "AC-070", "AC-071", "AC-072", "AC-073", "AC-074", "AC-080", "AC-081", "AC-082",
}

PARTIAL = {"AC-032", "AC-042", "AC-052", "AC-090"}
DEFER = {"AC-033", "AC-034", "AC-091"}

DONE_FILL = PatternFill("solid", fgColor="C6EFCE")
PARTIAL_FILL = PatternFill("solid", fgColor="FFE699")


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Backlog"]
    for row in ws.iter_rows(min_row=2):
        uid = str(row[0].value or "").strip()
        if uid in DONE:
            row[11].value = "Done"
            row[11].fill = DONE_FILL
        elif uid in PARTIAL:
            row[11].value = "Partial"
            row[11].fill = PARTIAL_FILL
        elif uid in DEFER:
            row[11].value = "Deferred"
    wb.save(XLSX)
    print(f"Updated {XLSX.name}: {len(DONE)} Done, {len(PARTIAL)} Partial, {len(DEFER)} Deferred")


if __name__ == "__main__":
    main()
